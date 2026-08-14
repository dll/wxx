#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_schedules.py — 教师/学生账号 + 课表批量导入（角色化）

数据源（放在 data/ 下）：
  - 教师.xlsx              教师账号清单（工号/姓名/...）；密码=工号
  - 194个教师课表.zip      《工号-姓名》课表.xlsx（含 课程/周次/节次/班级/地点）
  - 46个班级课表.zip       《班级》课表.xlsx（含 课程/周次；无地点）
  - 学生名单 xlsx          （计科251/252、2026级新生录取数据 等）：学号/姓名/班级

行为：
  1) 建/更新教师账号(username=工号, role=teacher, 密码=工号)，已存在跳过不改密码
  2) 解析教师课表 → 写 course_schedules(user_id=教师)
  3) 建/更新学生账号(username=学号, role=student, class_name=班级, 密码=学号)
  4) 解析班级课表 → 对有名单的班级，展开到每位学生 user_id

准确性：只从 xlsx 读账号/姓名；周次解析失败则跳过并计数；location 没有则留空(不编造)。
用法：python server/scripts/import_schedules.py --db <path> [--dry]
"""
import sys, io, os, re, zipfile, argparse, tempfile, hashlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import sqlite3
import openpyxl
import bcrypt

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")
DATA = os.path.join(BASE, "data")
TEACHER_XLSX = os.path.join(DATA, "教师.xlsx")
TEACHER_ZIP = os.path.join(DATA, "194个教师课表.zip")
CLASS_ZIP = os.path.join(DATA, "46个班级课表.zip")

# 学生名单：(文件名, 学号列, 姓名列, 班级列)；班级列值即班级名，用于匹配班级课表 zip
STUDENT_ROSTERS = [
    ("计科251学生名单_150800448.xlsx", 0, 1, 4),
    ("计科252学生名单_150800450.xlsx", 0, 1, 4),
    ("2026级新生录取数据.xlsx", 0, 1, 4),
]

PERIOD_MAP = {1: (8, 0), 2: (8, 55), 3: (10, 0), 4: (10, 55), 5: (14, 0),
              6: (14, 55), 7: (16, 0), 8: (16, 55), 9: (19, 0), 10: (19, 55)}


def bcrypt_pwd(p):
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt(rounds=10)).decode("utf-8")


def hash_match(db_hash, p):
    try:
        return bcrypt.checkpw(p.encode("utf-8"), db_hash.encode("utf-8"))
    except Exception:
        return False


# ---------- 课表格解析 ----------
def load_xlsx_bytes(zf, name):
    tmp = os.path.join(tempfile.gettempdir(), "wxx_" + hashlib.md5(name.encode()).hexdigest() + ".xlsx")
    with open(tmp, "wb") as f:
        f.write(zf.read(name))
    return tmp


def parse_timetable(path):
    """返回 [(weekday, start_period, end_period, course_name, location, teacher, weeks)]"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 10:
        return []
    out = []
    # 表头行：row2 = ["节次/星期","星期一"...]；weekday 1=周一
    header = [str(c) if c else "" for c in rows[1]]
    # 节次行：rows[2:] 对应 1-2节/3-4节/午间/5-6/7-8/晚间/9-10/11-12
    period_rows = [
        (2, (1, 2)), (3, (3, 4)),        # rows 3,4 -> 1-2节, 3-4节
        (5, (5, 6)), (6, (7, 8)),          # rows 6,7 -> 5-6,7-8
        (8, (9, 10)), (9, (11, 12)),
    ]
    for ridx, (sp, ep) in period_rows:
        if ridx >= len(rows):
            continue
        row = rows[ridx]
        for col in range(1, 8):  # 周一~周日
            cell = row[col] if col < len(row) else None
            if cell is None or str(cell).strip() == "":
                continue
            txt = str(cell).strip()
            # 提取 课程名/周次/地点/教师/班级（富文本格式，用分隔符拆分）
            parsed = parse_cell(txt, sp, ep)
            if parsed is None:
                continue
            for item in parsed:
                # (weekday, start, end, cname, location, teacher, weeks)
                out.append((col, item[0], item[1], item[2], item[3], item[4], item[5]))
    return out


def normalize_weeks(ws):
    """把'5-9' '2-5,8' '1-17周' '5-17单周' '2-4单周' 等转成标准 weeks_pattern"""
    ws = ws.replace("周", "").replace("节", "").strip().rstrip(",").rstrip("，")
    if not ws:
        return ""
    # 处理"单周/双周"
    parity = ""
    if re.search(r"单", ws):
        parity = "odd"
    elif re.search(r"双", ws):
        parity = "even"
    ws = re.sub(r"[单双周]", "", ws)
    # 逗号分隔段
    segs = [s.strip() for s in re.split(r"[,\uff0c;；]", ws) if s.strip()]
    out = []
    for seg in segs:
        m = re.match(r"^(\d+)-(\d+)$", seg)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            out.append(f"{start}-{end}" + (":odd" if parity == "odd" else ":even" if parity == "even" else ""))
        elif re.match(r"^\d+$", seg):
            n = int(seg)
            out.append(f"{n}-{n}" + (":odd" if parity == "odd" else ":even" if parity == "even" else ""))
    return ",".join(out)


def parse_cell(txt, sp, ep):
    """解析一个课表格单元格，可能含多节课（分段）。
    实测数据用 ◇(U+25C7)/◆(U+25C6) 作为统一分隔符切段：
      课程名 ◇ 周次[节次] ◇ 班级 ◇ 教师 ◇ 地点
    例：★数字逻辑◇1-16周[1-2节]◇计科251◇董再鹏,皮新语◇YF4206
    返回 [(start,end,cname,location,teacher,weeks)] 或 None。"""
    SEP = "[\u25c7\u25c6]"   # ◇◆
    results = []
    chunks = re.split(r"[\n；;]", txt)
    for ch in chunks:
        ch = ch.strip()
        if not ch or ch in ("午间", "晚间"):
            continue
        # 按 ◇ 切段
        segs = [s.strip() for s in re.split(SEP, ch) if s.strip()]
        if not segs:
            continue
        cname = re.sub(r"^[\u2605\u2606\u25c9★☆◉\s]*", "", segs[0]).strip()
        weeks = ""
        location = ""
        teacher = ""
        for seg in segs[1:]:
            # 周次段：含 数字/周/单双
            if re.match(r"^[\d\s,，-]+[单双周]*周?", seg) and not re.search(r"[\u4e00-\u9fff]{2,}", re.sub(r"[单双周末节周\[\]\s\d,-，]", "", seg)):
                mw = re.match(r"([\d\s,，-]+[单双周]*周?)", seg)
                if mw:
                    weeks = normalize_weeks(mw.group(1))
            # 地点段：字母数字 或 教? 或 含"室"
            elif re.search(r"[A-Za-z]\d+\w*|教\s*\d+|\d+室|楼", seg):
                location = seg
            # 教师段：纯中文2-4字(可能逗号多人 -> 全保留)
            elif re.match(r"^[\u4e00-\u9fff,，]{2,}$", seg):
                teacher = seg
        if not cname:
            continue
        results.append((sp, ep, cname, location, teacher, weeks))
    return results if results else None


# ---------- 主流程 ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default=None, help="SQLite 数据库路径（默认 server/data/wxx.db）")
    ap.add_argument("--dry", action="store_true", help="仅统计不写库")
    ap.add_argument("--students", type=str, default="计科251,计科252,计科261,计科262,软件261,软件262,网安26,空间26,大数据26",
                    help="要导入学生课表的班级（逗号分隔）")
    args = ap.parse_args()
    db = args.db or os.path.join(BASE, "server", "data", "wxx.db")
    conn = sqlite3.connect(db)
    cur = conn.cursor()

    report = {"teacher_acct_new": 0, "teacher_acct_skip": 0, "teacher_sched": 0,
              "student_acct_new": 0, "student_acct_skip": 0, "student_sched": 0,
              "skip_no_acct": 0, "skip_parse": 0}

    # ── 1) 教师账号：教师.xlsx ──
    print("== 1) 教师账号 ==")
    if os.path.exists(TEACHER_XLSX):
        wb = openpyxl.load_workbook(TEACHER_XLSX, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        teach_rows = [r for r in rows if r and r[0]]
        wb.close()
        for r in teach_rows:
            uname = str(r[0]).strip(); name = str(r[1]).strip() if len(r) > 1 else ""
            if not uname:
                continue
            exist = cur.execute("select id,password_hash from users where username=?", (uname,)).fetchone()
            if exist:
                report["teacher_acct_skip"] += 1
            else:
                pwd = bcrypt_pwd(uname)
                if not args.dry:
                    cur.execute("INSERT INTO users (username, display_name, role, owner_scope, owner_id, password_hash, status) VALUES (?,?,?,?,?,?, 'active')",
                                (uname, name, "teacher", "college", "cs", pwd))
                report["teacher_acct_new"] += 1
        if not args.dry:
            conn.commit()
        print(f"  教师.xlsx 共 {len(teach_rows)} 位：新增 {report['teacher_acct_new']}，已存在跳过 {report['teacher_acct_skip']}")

    # ── 2) 教师课表 ──
    print("== 2) 教师课表 ==")
    if os.path.exists(TEACHER_ZIP):
        zf = zipfile.ZipFile(TEACHER_ZIP)
        for n in zf.namelist():
            m = re.match(r"《(.+?)-", n)
            if not m:
                continue
            uname = m.group(1).strip()
            uid = cur.execute("select id from users where username=?", (uname,)).fetchone()
            tmp = load_xlsx_bytes(zf, n)
            try:
                rows = parse_timetable(tmp)
            except Exception as e:
                report["skip_parse"] += 1
                continue
            if not rows:
                report["skip_parse"] += 1
                continue
            if not uid:
                report["skip_no_acct"] += len(rows)
                print(f"  [{uname}] 无账号，跳过 {len(rows)} 条")
                continue
            uid = uid[0]
            for (wd, sp, ep, cname, loc, teacher, weeks) in rows:
                cid = hashlib.md5(f"{uname}-{cname}-{wd}-{sp}".encode()).hexdigest()[:16]
                if not args.dry:
                    cur.execute(
                        "INSERT OR IGNORE INTO course_schedules (user_id, course_id, course_name, semester_code, weekday, start_period, end_period, weeks_pattern, location, teacher) VALUES (?,?,?,?,?,?,?,?,?,?)",
                        (uid, cid, cname, "2026-2027-1", wd, sp, ep, weeks or "", loc, teacher or ""))
                report["teacher_sched"] += 1
        zf.close()
        if not args.dry:
            conn.commit()
        print(f"  教师课表导入 {report['teacher_sched']} 条")

    # ── 3) 学生账号：名单 ──
    print("== 3) 学生账号 ==")
    for fn, idc, nmc, clsc in STUDENT_ROSTERS:
        p = os.path.join(DATA, fn)
        if not os.path.exists(p):
            print(f"  缺名单: {fn}")
            continue
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # header
        cnt_new = cnt_skip = 0
        for r in rows:
            if not r or not r[idc]:
                continue
            uname = str(r[idc]).strip()
            name = str(r[nmc]).strip() if len(r) > nmc else ""
            clsname = str(r[clsc]).strip() if len(r) > clsc else ""
            exist = cur.execute("select id from users where username=?", (uname,)).fetchone()
            if exist:
                cnt_skip += 1
            else:
                pwd = bcrypt_pwd(uname)
                if not args.dry:
                    cur.execute("INSERT INTO users (username, display_name, role, owner_scope, owner_id, password_hash, status, class_name, major) VALUES (?,?,?,?,?,?, 'active', ?, ?)",
                                (uname, name, "student", "college", "cs", pwd, clsname, clsname))
                cnt_new += 1
        wb.close()
        if not args.dry:
            conn.commit()
        report["student_acct_new"] += cnt_new
        report["student_acct_skip"] += cnt_skip
        print(f"  {fn}: 新增 {cnt_new}，已存在 {cnt_skip}")

    # ── 4) 班级课表：仅导入 --students 指定班级 ──
    print("== 4) 班级课表 ==")
    want_classes = set(c.strip() for c in args.students.split(",") if c.strip())
    if os.path.exists(CLASS_ZIP):
        zf = zipfile.ZipFile(CLASS_ZIP)
        for n in zf.namelist():
            m = re.match(r"《(.+?)》课表", n)
            if not m:
                continue
            clsname = m.group(1).strip()
            if clsname not in want_classes:
                print(f"  [{clsname}] 不在导入范围，跳过")
                continue
            # 该班学生 user_id
            stu_ids = [r[0] for r in cur.execute(
                "select id from users where role='student' and class_name=?", (clsname,))]
            if not stu_ids:
                report["skip_no_acct"] += 1
                print(f"  [{clsname}] 无可导入学生账号，跳过")
                continue
            tmp = load_xlsx_bytes(zf, n)
            try:
                rows = parse_timetable(tmp)
            except Exception as e:
                report["skip_parse"] += 1
                continue
            for uid in stu_ids:
                for (wd, sp, ep, cname, loc, teacher, weeks) in rows:
                    cid = hashlib.md5(f"{clsname}-{cname}-{wd}-{sp}".encode()).hexdigest()[:16]
                    if not args.dry:
                        cur.execute(
                            "INSERT OR IGNORE INTO course_schedules (user_id, course_id, course_name, semester_code, weekday, start_period, end_period, weeks_pattern, location, teacher) VALUES (?,?,?,?,?,?,?,?,?,?)",
                            (uid, cid, cname, "2026-2027-1", wd, sp, ep, weeks or "", loc or "", teacher or ""))
                    report["student_sched"] += 1
            print(f"  [{clsname}] 学生 {len(stu_ids)} 人，课表 {len(rows)} 节课，展开 {len(rows)*len(stu_ids)} 条")
        zf.close()
        if not args.dry:
            conn.commit()

    if not args.dry:
        conn.commit()
    conn.close()
    print("\n===== 导入报告 =====")
    for k, v in report.items():
        print(f"  {k}: {v}")
    print("\n(若 --dry 未写库；去掉 --dry 执行真实导入。默认密码=工号/学号。)")


if __name__ == "__main__":
    main()

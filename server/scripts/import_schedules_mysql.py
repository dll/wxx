#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_schedules_mysql.py — 真实课表/教师/学生账号批量导入（MySQL 生产版）

数据源（与 SQLite 版 import_schedules.py 相同，放在仓库 data/ 下，.gitignore 已排除不入 git）：
  - 教师.xlsx              教师账号清单（工号/姓名/...）；密码=工号
  - 194个教师课表.zip      《工号-姓名》课表.xlsx
  - 46个班级课表.zip       《班级》课表.xlsx
  - 学生名单 xlsx          计科251/252、2026级新生录取、roster_templates/

与 SQLite 版差异：连接 MySQL（-h127.0.0.1），INSERT 用 INSERT IGNORE，补充 college 字段。
只读真实 xlsx/zip，周次解析失败则跳过，不编造。支持 --dry 仅统计不写库。
用法：
  python server/scripts/import_schedules_mysql.py --dry [--students 计科251,计科252] [--owner cs]
  （生产真实导入：不带 --dry；DB 密码自动读 /etc/wxx/env 的 DB_PASSWORD，或 --db-pass 传入）
"""
import sys, io, os, re, zipfile, argparse, tempfile, hashlib
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pymysql
import openpyxl
import bcrypt

BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")
DATA = os.path.join(BASE, "data")
TEACHER_XLSX = os.path.join(DATA, "教师.xlsx")
TEACHER_ZIP = os.path.join(DATA, "194个教师课表.zip")
CLASS_ZIP = os.path.join(DATA, "46个班级课表.zip")

STUDENT_ROSTERS = [
    ("计科251学生名单_150800448.xlsx", 0, 1, 4),
    ("计科252学生名单_150800450.xlsx", 0, 1, 4),
    ("2026级新生录取数据.xlsx", 0, 1, 4),
]


def discover_roster_files():
    found = []
    for d in [DATA, os.path.join(DATA, "roster_templates")]:
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            m = re.match(r"《(.+?)》学生名单\.xlsx$", fn)
            if m:
                if any(fn == s[0] for s in STUDENT_ROSTERS):
                    continue
                found.append((m.group(1), os.path.join(d, fn), 0, 1, 4))
    return found


def bcrypt_pwd(p):
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt(rounds=10)).decode("utf-8")


def load_xlsx_bytes(zf, name):
    tmp = os.path.join(tempfile.gettempdir(), "wxx_" + hashlib.md5(name.encode()).hexdigest() + ".xlsx")
    with open(tmp, "wb") as f:
        f.write(zf.read(name))
    return tmp


def parse_timetable(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 10:
        return []
    out = []
    period_rows = [
        (2, (1, 2)), (3, (3, 4)),
        (5, (5, 6)), (6, (7, 8)),
        (8, (9, 10)), (9, (11, 12)),
    ]
    for ridx, (sp, ep) in period_rows:
        if ridx >= len(rows):
            continue
        row = rows[ridx]
        for col in range(1, 8):
            cell = row[col] if col < len(row) else None
            if cell is None or str(cell).strip() == "":
                continue
            txt = str(cell).strip()
            parsed = parse_cell(txt, sp, ep)
            if parsed is None:
                continue
            for item in parsed:
                out.append((col, item[0], item[1], item[2], item[3], item[4], item[5]))
    return out


def normalize_weeks(ws):
    ws = ws.replace("周", "").replace("节", "").strip().rstrip(",").rstrip("，")
    if not ws:
        return ""
    parity = ""
    if re.search(r"单", ws):
        parity = "odd"
    elif re.search(r"双", ws):
        parity = "even"
    ws = re.sub(r"[单双周]", "", ws)
    segs = [s.strip() for s in re.split(r"[,\uff0c;；]", ws) if s.strip()]
    out = []
    for seg in segs:
        m = re.match(r"^(\d+)-(\d+)$", seg)
        if m:
            start, end = int(m.group(1)), int(m.group(2))
            out.append(f"{start}-{end}" + (":odd" if parity == "odd" else ":even" if parity == "even" else ""))
        elif re.match(r"^\d+$", seg):
            n = int(seg)
            out.append(f"{n}-{n}" + (":odd" if parity == "odd" else ":even" if parity == "even" else ""))
    return ",".join(out)


def parse_cell(txt, sp, ep):
    SEP = "[\u25c7\u25c6]"
    results = []
    chunks = re.split(r"[\n；;]", txt)
    for ch in chunks:
        ch = ch.strip()
        if not ch or ch in ("午间", "晚间"):
            continue
        segs = [s.strip() for s in re.split(SEP, ch) if s.strip()]
        if not segs:
            continue
        cname = re.sub(r"^[\u2605\u2606\u25c9★☆◉\s]*", "", segs[0]).strip()
        weeks = ""
        location = ""
        teacher = ""
        for seg in segs[1:]:
            if re.match(r"^[\d\s,，-]+[单双周]*周?", seg) and not re.search(r"[\u4e00-\u9fff]{2,}", re.sub(r"[单双周末节周\[\]\s\d,-，]", "", seg)):
                mw = re.match(r"([\d\s,，-]+[单双周]*周?)", seg)
                if mw:
                    weeks = normalize_weeks(mw.group(1))
            elif re.search(r"[A-Za-z]\d+\w*|教\s*\d+|\d+室|楼", seg):
                location = seg
            elif re.match(r"^[\u4e00-\u9fff,，]{2,}$", seg):
                teacher = seg
        if not cname:
            continue
        results.append((sp, ep, cname, location, teacher, weeks))
    return results if results else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry", action="store_true", help="仅统计不写库")
    ap.add_argument("--students", type=str,
                    default="计科251,计科252,计科261,计科262,软件261,软件262,网安26,空间26,大数据26",
                    help="要导入学生课表的班级（逗号分隔）")
    ap.add_argument("--semester", type=str, default="2026-2027-1")
    ap.add_argument("--owner", type=str, default="cs", help="学院 owner_id（默认 cs=计算机学院）")
    ap.add_argument("--db-host", default="127.0.0.1")
    ap.add_argument("--db-user", default="wxx")
    ap.add_argument("--db-pass", default=None, help="默认从 /etc/wxx/env 读 DB_PASSWORD")
    ap.add_argument("--db-name", default="wxx")
    args = ap.parse_args()

    semester_code = args.semester
    owner_id = args.owner

    db_pass = args.db_pass
    if db_pass is None and os.path.exists("/etc/wxx/env"):
        for line in open("/etc/wxx/env", encoding="utf-8", errors="ignore"):
            if line.strip().startswith("DB_PASSWORD="):
                db_pass = line.strip().split("=", 1)[1]
    if not db_pass:
        print("[FATAL] 未提供 DB 密码（--db-pass 或 /etc/wxx/env 的 DB_PASSWORD）")
        sys.exit(1)

    conn = pymysql.connect(host=args.db_host, user=args.db_user, password=db_pass,
                           database=args.db_name, charset="utf8mb4")
    cur = conn.cursor()

    report = {"teacher_acct_new": 0, "teacher_acct_skip": 0, "teacher_sched": 0,
              "student_acct_new": 0, "student_acct_skip": 0, "student_sched": 0,
              "skip_no_acct": 0, "skip_parse": 0}

    # ── 1) 教师账号 ──
    print("== 1) 教师账号 ==")
    if os.path.exists(TEACHER_XLSX):
        wb = openpyxl.load_workbook(TEACHER_XLSX, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        teach_rows = [r for r in rows if r and r[0]]
        wb.close()
        for r in teach_rows:
            uname = str(r[0]).strip(); name = str(r[1]).strip() if len(r) > 1 else ""
            if not uname:
                continue
            cur.execute("SELECT id FROM users WHERE username=%s", (uname,))
            if cur.fetchone():
                report["teacher_acct_skip"] += 1
            else:
                pwd = bcrypt_pwd(uname)
                if not args.dry:
                    cur.execute(
                        "INSERT IGNORE INTO users (username, display_name, role, owner_scope, owner_id, password_hash, status, college, major, class_name) VALUES (%s,%s,%s,%s,%s,%s,'active',%s,%s,%s)",
                        (uname, name, "teacher", "college", owner_id, pwd, "", "", ""))
                report["teacher_acct_new"] += 1
        if not args.dry:
            conn.commit()
        print(f"  教师.xlsx 共 {len(teach_rows)} 位：新增 {report['teacher_acct_new']}，已存在跳过 {report['teacher_acct_skip']}")

    # ── 2) 教师课表 ──
    print("== 2) 教师课表 ==")
    if os.path.exists(TEACHER_ZIP):
        zf = zipfile.ZipFile(TEACHER_ZIP)
        for n in zf.namelist():
            m = re.match(r"《(.+?)-", n)
            if not m:
                continue
            uname = m.group(1).strip()
            cur.execute("SELECT id FROM users WHERE username=%s", (uname,))
            uid = cur.fetchone()
            tmp = load_xlsx_bytes(zf, n)
            try:
                rows = parse_timetable(tmp)
            except Exception:
                report["skip_parse"] += 1
                continue
            if not rows:
                report["skip_parse"] += 1
                continue
            if not uid:
                report["skip_no_acct"] += len(rows)
                print(f"  [{uname}] 无账号，跳过 {len(rows)} 条")
                continue
            uid = uid[0]
            for (wd, sp, ep, cname, loc, teacher, weeks) in rows:
                cid = hashlib.md5(f"{uname}-{cname}-{wd}-{sp}".encode()).hexdigest()[:16]
                if not args.dry:
                    cur.execute(
                        "INSERT IGNORE INTO course_schedules (user_id, course_id, course_name, semester_code, weekday, start_period, end_period, weeks_pattern, location, teacher) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (uid, cid, cname, semester_code, wd, sp, ep, weeks or "", loc or "", teacher or ""))
                report["teacher_sched"] += 1
        zf.close()
        if not args.dry:
            conn.commit()
        print(f"  教师课表导入 {report['teacher_sched']} 条")

    # ── 3) 学生账号 ──
    print("== 3) 学生账号 ==")
    roster_files = []
    for item in STUDENT_ROSTERS:
        fn, idc, nmc, clsc = item
        roster_files.append((fn, os.path.join(DATA, fn), idc, nmc, clsc, None))
    for cls_name, tpl_path, idc, nmc, clsc in discover_roster_files():
        roster_files.append((os.path.basename(tpl_path), tpl_path, idc, nmc, clsc, cls_name))
    template_class_names = set(x[5] for x in roster_files if x[5])

    for fn, p, idc, nmc, clsc, cls_from_name in roster_files:
        if not os.path.exists(p):
            print(f"  缺名单: {os.path.basename(p)}")
            continue
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        next(rows, None)
        cnt_new = cnt_skip = 0
        for r in rows:
            if not r or not r[idc]:
                continue
            uname = str(r[idc]).strip()
            name = str(r[nmc]).strip() if len(r) > nmc else ""
            clsname = (cls_from_name or (str(r[clsc]).strip() if len(r) > clsc else "")).strip()
            cur.execute("SELECT id FROM users WHERE username=%s", (uname,))
            if cur.fetchone():
                cnt_skip += 1
            else:
                pwd = bcrypt_pwd(uname)
                if not args.dry:
                    cur.execute(
                        "INSERT IGNORE INTO users (username, display_name, role, owner_scope, owner_id, password_hash, status, class_name, major) VALUES (%s,%s,%s,%s,%s,%s,'active',%s,%s)",
                        (uname, name, "student", "college", owner_id, pwd, clsname, clsname))
                cnt_new += 1
        wb.close()
        if not args.dry:
            conn.commit()
        print(f"  {os.path.basename(p)}: 新增 {cnt_new}，已存在 {cnt_skip}（班级={cls_from_name or '—'}）")

    # ── 4) 班级课表：按班级展开到学生 ──
    print("== 4) 班级课表 ==")
    want_classes = set(c.strip() for c in args.students.split(",") if c.strip())
    cur.execute("SELECT DISTINCT class_name FROM users WHERE role='student' AND class_name != ''")
    want_classes |= set(str(r[0]) for r in cur.fetchall())
    template_has = set(x[5] for x in roster_files if x[5])
    want_classes |= template_has
    if os.path.exists(CLASS_ZIP):
        zf = zipfile.ZipFile(CLASS_ZIP)
        for n in zf.namelist():
            m = re.match(r"《(.+?)》课表", n)
            if not m:
                continue
            clsname = m.group(1).strip()
            if clsname not in want_classes:
                print(f"  [{clsname}] 不在导入范围，跳过")
                continue
            cur.execute("SELECT id FROM users WHERE role='student' AND class_name=%s", (clsname,))
            stu_ids = [r[0] for r in cur.fetchall()]
            if not stu_ids:
                report["skip_no_acct"] += 1
                print(f"  [{clsname}] 无可导入学生账号，跳过")
                continue
            tmp = load_xlsx_bytes(zf, n)
            try:
                rows = parse_timetable(tmp)
            except Exception:
                report["skip_parse"] += 1
                continue
            for uid in stu_ids:
                for (wd, sp, ep, cname, loc, teacher, weeks) in rows:
                    cid = hashlib.md5(f"{clsname}-{cname}-{wd}-{sp}".encode()).hexdigest()[:16]
                    if not args.dry:
                        cur.execute(
                            "INSERT IGNORE INTO course_schedules (user_id, course_id, course_name, semester_code, weekday, start_period, end_period, weeks_pattern, location, teacher) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                            (uid, cid, cname, semester_code, wd, sp, ep, weeks or "", loc or "", teacher or ""))
                    report["student_sched"] += 1
            print(f"  [{clsname}] 学生 {len(stu_ids)} 人，课表 {len(rows)} 节课，展开 {len(rows)*len(stu_ids)} 条")
        zf.close()
        if not args.dry:
            conn.commit()

    if not args.dry:
        conn.commit()
    conn.close()
    print("\n===== 导入报告 =====")
    for k, v in report.items():
        print(f"  {k}: {v}")
    print("\n(若 --dry 未写库；去掉 --dry 执行真实导入。默认密码=工号/学号。)")


if __name__ == "__main__":
    main()
